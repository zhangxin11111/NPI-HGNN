from openpyxl import load_workbook
import math
import pandas as pd
import networkx as nx
import numpy as np
import random
import os
import os.path as osp
from node2vec import Node2Vec
def read_node_dict_file(path):
    node_dict={}
    with open(path,mode='r') as f:
        lines=f.readlines()
        for line in lines:
            line_list=line.strip().split(',')
            node_dict[line_list[0]]=line_list[1]
    return node_dict
def get_subgraph(protein_set,interactions):
    subgraph=set()
    for i in protein_set:
        for j in protein_set:
            if (i,j) in interactions:
                subgraph.add((i,j))
            elif (j,i) in interactions:
                subgraph.add((j,i))
    return subgraph
def read_RPI(path):
    if not osp.exists(path):
        print('RPI文件不存在')
        return set()
    print('读取RPI文件')
    wb = load_workbook(path)
    sheets = wb.worksheets
    sheet = sheets[0]
    rows = sheet.rows
    flag = 0
    ppi = set()
    for row in rows:
        if flag == 0:
            flag = flag + 1
            continue
        node1_name=row[0].value
        node2_name=row[1].value
        ppi.add((node1_name,node2_name))
    return ppi
def read_PPI(path):
    if not osp.exists(path):
        return set()
    wb = load_workbook(path)
    sheets = wb.worksheets
    sheet = sheets[0]
    rows = sheet.rows
    flag = 0
    ppi = set()
    for row in rows:
        if flag == 0:
            flag = flag + 1
            continue
        node1_name=row[0].value
        node2_name=row[1].value
        ppi.add((node1_name,node2_name))
    return ppi
def read_RRI(path):
    if not osp.exists(path):
        return set()
    wb = load_workbook(path)
    sheets = wb.worksheets
    sheet = sheets[0]
    rows = sheet.rows
    flag = 0
    rri = set()
    for row in rows:
        if flag == 0:
            flag = flag + 1
            continue
        node1_name=row[0].value
        node2_name=row[1].value
        rri.add((node1_name,node2_name))
    return rri
def read_RRI_file(path):
    npi_df = pd.read_excel(path)
    npi_df = npi_df.rename(columns={'RNA1 names': 'source', 'RNA2 names': 'target', 'Labels': 'scores'})
    npi_df = npi_df[['source', 'target', 'scores']]
    return npi_df
def read_RPI_file(path):
    npi_df = pd.read_excel(path)
    npi_df = npi_df.rename(columns={'RNA names': 'source', 'Protein names': 'target', 'Labels': 'type'})
    npi_df['type'] = 0
    npi_df = npi_df[['source', 'target', 'type']]
    return npi_df
def read_all(path,projectName):
    edge_list = pd.read_csv(path, header=None).reset_index(drop=True)
    edge_list=np.array(edge_list)
    rna_names=set()
    protein_names=set()
    G = nx.Graph()
    for edge in edge_list:
        rna_names.add(edge[0])
        protein_names.add(edge[1])
        G.add_edge(edge[0], edge[1])
    ppi = read_PPI(f'../../data/{projectName}/processed_database_data/{projectName}_PPI.xlsx')
    for i in ppi:
        protein_names.add(i[0])
        protein_names.add(i[1])
        G.add_edge(i[0],i[1])
    rri = read_RRI(f'../../data/{projectName}/processed_database_data/{projectName}_RRI.xlsx')
    for i in rri:
        rna_names.add(i[0])
        rna_names.add(i[1])
        G.add_edge(i[0],i[1])
    G = G.to_undirected()
    return G,rna_names,protein_names
def read_rpin(path):
    edge_list = pd.read_csv(path, header=None).reset_index(drop=True)
    edge_list=np.array(edge_list)
    rna_names=set()
    protein_names=set()
    G = nx.Graph()
    for edge in edge_list:
        rna_names.add(edge[0])
        protein_names.add(edge[1])
        G.add_edge(edge[0], edge[1])
    G = G.to_undirected()
    return G,rna_names,protein_names
def read_ppin(path):
    protein_names=set()
    G = nx.Graph()
    ppi = read_PPI(path)
    for i in ppi:
        protein_names.add(i[0])
        protein_names.add(i[1])
        G.add_edge(i[0],i[1])
    G = G.to_undirected()
    return G,protein_names
def read_rrsn(path):
    rna_names=set()
    G = nx.Graph()
    rri = read_RRI(path)
    for i in rri:
        rna_names.add(i[0])
        rna_names.add(i[1])
        G.add_edge(i[0],i[1])
    G = G.to_undirected()
    return G,rna_names
def read_rpi(npi_path):
    npi_df = pd.read_excel(npi_path)
    npi_df = npi_df.rename(columns={'RNA names': 'source', 'Protein names': 'target', 'Labels': 'type'})
    npi_df['type'] = 0
    rna_name_set=set(npi_df['source'].tolist())
    protein_name_set=set(npi_df['target'].tolist())
    npi_df = npi_df[['source', 'type','target']]
    from sklearn.utils import shuffle
    npi_df = shuffle(npi_df)
    return npi_df,rna_name_set,protein_name_set
def random_negative_sampling(set_interaction,rna_name_set,protein_name_set, size):
    set_negativeInteraction = set()
    num_of_ncRNA = len(rna_name_set)
    rna_name_list=list(rna_name_set)
    protein_name_list = list(protein_name_set)
    num_of_protein = len(protein_name_set)
    negative_interaction_count = 0
    while(negative_interaction_count < size):
        random_index_ncRNA = random.randint(0, num_of_ncRNA - 1)
        random_index_protein = random.randint(0, num_of_protein - 1)
        temp_ncRNA = rna_name_list[random_index_ncRNA]
        temp_protein = protein_name_list[random_index_protein]
        negativeInteraction = (temp_ncRNA, temp_protein)
        if negativeInteraction in set_interaction:
            continue
        if negativeInteraction in set_negativeInteraction:
            continue
        set_negativeInteraction.add(negativeInteraction)
        negative_interaction_count = negative_interaction_count + 1

    return set_negativeInteraction
def write_interactor(interactor,output_path):
    with open(output_path,'w') as f:
        for triplet in interactor:
            f.write(triplet[0]+','+triplet[1]+'\n')
def generate_n2v(G,path,dimensions, walk_length, num_walks,p, q, workers):
    node2vec = Node2Vec(G, dimensions=dimensions, walk_length=walk_length, num_walks=num_walks,p=p, q=q, workers=workers)
    model = node2vec.fit()
    if not osp.exists(path):
        os.makedirs(path)
    model.wv.save_word2vec_format(path + '/result.emb')
def generate_node_vec_with_fre(fold,projectName,node_path):
    node_name_vec_dict=read_node2vec_file(f'{node_path}/node2vec/{fold}/result.emb')
    input_path_rna = f'../../data/{projectName}/processed_database_data/ncRNA_sequence.fasta'
    rna_name_list, rna_sequence_list = read_sequence_file(path=input_path_rna)
    rna_kmer_fre_path=f'../../data/{projectName}/frequency/rna_seq/result.emb'
    fre_name_list, fre_list=read_kmer_fre_file(rna_kmer_fre_path)
    fre_name_index_dict = {}
    for index in range(len(fre_name_list)):
        fre_name_index_dict[fre_name_list[index]] = index
    if not osp.exists(f'{node_path}/node_vec/{fold}'):
        os.makedirs(f'{node_path}/node_vec/{fold}')
    with open(f'{node_path}/node_vec/{fold}/rna_vec.txt',mode='w') as f:
        for rna_name in rna_name_list:
            node_vec=fre_list[fre_name_index_dict[rna_name]]
            for i in range(49):
                node_vec.append('0')
            node2vec=node_name_vec_dict[rna_name]
            node_vec.extend(node2vec)
            f.write(rna_name+','+','.join(node_vec)+'\n')
    input_path_protein = f'../../data/{projectName}/processed_database_data/protein_sequence.fasta'
    protein_name_list, protein_sequence_list = read_sequence_file(path=input_path_protein)
    protein_kmer_fre_path=f'../../data/{projectName}/frequency/protein_seq/result.emb'
    fre_name_list, fre_list=read_kmer_fre_file(protein_kmer_fre_path)
    fre_name_index_dict = {}
    for index in range(len(fre_name_list)):
        fre_name_index_dict[fre_name_list[index]] = index
    with open(f'{node_path}/node_vec/{fold}/protein_vec.txt',mode='w') as f:
        for protein_name in protein_name_list:
            node_vec=[]
            for i in range(64):
                node_vec.append('0')
            node_vec.extend(fre_list[fre_name_index_dict[protein_name]])
            node2vec=node_name_vec_dict[protein_name]
            node_vec.extend(node2vec)
            f.write(protein_name+','+','.join(node_vec)+'\n')
def generate_node_vec_with_pyfeat(fold,projectName,node_path):
    node_name_vec_dict=read_node2vec_file(f'{node_path}/node2vec/{fold}/result.emb')
    input_path_rna = f'../../data/{projectName}/processed_database_data/ncRNA_sequence.fasta'
    rna_name_list, rna_sequence_list = read_sequence_file(path=input_path_rna)
    rna_pyfeat_path=f'../../data/{projectName}/pyfeat/rna_seq/result.emb'
    pyfeat_name_list, pyfeat_list=read_pyfeat_file(rna_pyfeat_path)
    pyfeat_name_index_dict = {} #rna_name,id
    for index in range(len(pyfeat_name_list)):
        pyfeat_name_index_dict[pyfeat_name_list[index]] = index
    if not osp.exists(f'{node_path}/node_vec/{fold}'):
        os.makedirs(f'{node_path}/node_vec/{fold}')
    with open(f'{node_path}/node_vec/{fold}/rna_vec.txt',mode='w') as f:
        for rna_name in rna_name_list:
            node_vec=pyfeat_list[pyfeat_name_index_dict[rna_name]]
            for i in range(100):
                node_vec.append('0')
            node2vec=node_name_vec_dict[rna_name]
            node_vec.extend(node2vec)
            f.write(rna_name+','+','.join(node_vec)+'\n')
    input_path_protein = f'../../data/{projectName}/processed_database_data/protein_sequence.fasta'
    protein_name_list, protein_sequence_list = read_sequence_file(path=input_path_protein)
    protein_pyfeat_path=f'../../data/{projectName}/pyfeat/protein_seq/result.emb'
    pyfeat_name_list, pyfeat_list=read_pyfeat_file(protein_pyfeat_path)
    pyfeat_name_index_dict = {} #protein_name,id
    for index in range(len(pyfeat_name_list)):
        pyfeat_name_index_dict[pyfeat_name_list[index]] = index
    with open(f'{node_path}/node_vec/{fold}/protein_vec.txt',mode='w') as f:
        for protein_name in protein_name_list:
            node_vec=[]
            for i in range(100):
                node_vec.append('0')
            node_vec.extend(pyfeat_list[pyfeat_name_index_dict[protein_name]])
            node2vec=node_name_vec_dict[protein_name]
            node_vec.extend(node2vec)
            f.write(protein_name+','+','.join(node_vec)+'\n')
def generate_node_vec_only_n2v(fold,projectName,node_path):
    node_name_vec_dict=read_node2vec_file(f'{node_path}/node2vec/{fold}/result.emb')
    input_path_rna = f'../../data/{projectName}/processed_database_data/ncRNA_sequence.fasta'
    rna_name_list, rna_sequence_list = read_sequence_file(path=input_path_rna)
    if not osp.exists(f'{node_path}/node_vec/{fold}'):
        os.makedirs(f'{node_path}/node_vec/{fold}')
    with open(f'{node_path}/node_vec/{fold}/rna_vec.txt',mode='w') as f:
        for rna_name in rna_name_list:
            node2vec=node_name_vec_dict[rna_name]
            f.write(rna_name+','+','.join(node2vec)+'\n')
    input_path_protein = f'../../data/{projectName}/processed_database_data/protein_sequence.fasta'
    protein_name_list, protein_sequence_list = read_sequence_file(path=input_path_protein)
    with open(f'{node_path}/node_vec/{fold}/protein_vec.txt',mode='w') as f:
        for protein_name in protein_name_list:
            node2vec=node_name_vec_dict[protein_name]
            f.write(protein_name+','+','.join(node2vec)+'\n')
def generate_node_vec_only_frequency(fold,projectName,node_path):
    input_path_rna = f'../../data/{projectName}/processed_database_data/ncRNA_sequence.fasta'
    rna_name_list, rna_sequence_list = read_sequence_file(path=input_path_rna)
    rna_kmer_fre_path=f'../../data/{projectName}/frequency/rna_seq/result.emb'
    fre_name_list, fre_list=read_kmer_fre_file(rna_kmer_fre_path)
    fre_name_index_dict = {} #rna_name,id
    for index in range(len(fre_name_list)):
        fre_name_index_dict[fre_name_list[index]] = index
    if not osp.exists(f'{node_path}/node_vec/{fold}'):
        os.makedirs(f'{node_path}/node_vec/{fold}')
    with open(f'{node_path}/node_vec/{fold}/rna_vec.txt',mode='w') as f:
        for rna_name in rna_name_list:
            node_vec=fre_list[fre_name_index_dict[rna_name]]
            for i in range(49):
                node_vec.append('0')
            f.write(rna_name+','+','.join(node_vec)+'\n')
    input_path_protein = f'../../data/{projectName}/processed_database_data/protein_sequence.fasta'
    protein_name_list, protein_sequence_list = read_sequence_file(path=input_path_protein)
    protein_kmer_fre_path=f'../../data/{projectName}/frequency/protein_seq/result.emb'
    fre_name_list, fre_list=read_kmer_fre_file(protein_kmer_fre_path)
    fre_name_index_dict = {} #protein_name,id
    for index in range(len(fre_name_list)):
        fre_name_index_dict[fre_name_list[index]] = index
    with open(f'{node_path}/node_vec/{fold}/protein_vec.txt',mode='w') as f:
        for protein_name in protein_name_list:
            node_vec=[]
            for i in range(64):
                node_vec.append('0')
            node_vec.extend(fre_list[fre_name_index_dict[protein_name]])
            f.write(protein_name+','+','.join(node_vec)+'\n')
def read_sequence_file(path):
    name_list = []
    sequence_list = []
    sequence_file_path = path
    sequence_file = open(sequence_file_path, mode='r')

    for line in sequence_file.readlines():
        if line[0].strip()=='':
            continue
        if line[0] == '>':
            sequence_name = line.strip()[1:]
            name_list.append(sequence_name)
        else:
            sequence_list.append(line.strip())
    sequence_file.close()
    return name_list, sequence_list
def read_kmer_fre_file(path):
    name_list = []
    fre_list = []
    kmer_fre_path = path
    kmer_fre_file = open(kmer_fre_path, mode='r')

    lines = kmer_fre_file.readlines()
    for i in range(len(lines)):
        line = lines[i].strip().split(',')
        seq_name = line[0]
        name_list.append(seq_name)
        seq_vector = line[1:]
        fre_list.append(seq_vector)
    kmer_fre_file.close()
    return name_list, fre_list
def read_pyfeat_file(path):
    name_list = []
    pyfeat_list = []
    pyfeat_file = open(path, mode='r')

    lines = pyfeat_file.readlines()
    for i in range(len(lines)):
        line = lines[i].strip().split(',')
        seq_name = line[0]
        name_list.append(seq_name)
        seq_vector = line[1:]
        pyfeat_list.append(seq_vector)
    pyfeat_file.close()
    return name_list, pyfeat_list
def read_node2vec_file(path):
    node_name_vec_dict={}
    node2vec_file = open(path, mode='r')
    lines = node2vec_file.readlines()
    for i in range(len(lines)):
        line = lines[i].strip().split(' ')
        node_name_vec_dict[line[0]]=line[1:]
    node2vec_file.close()
    return node_name_vec_dict
def generate_dataset_path(projectName,fold,nodeVecType,subgraph_type,samplingType,type):
    if samplingType==0: #    random fire
        if nodeVecType == 0:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/frequency/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/pyfeat/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/pyfeat/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/pyfeat/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_n2v/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_n2v/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_n2v/dataset/only_rpin/dataset_{fold}/{type}'
        else:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/random/only_frequency/dataset/only_rpin/dataset_{fold}/{type}'
    elif samplingType==1:
        if nodeVecType == 0:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/frequency/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/pyfeat/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/pyfeat/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/pyfeat/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_n2v/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_n2v/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_n2v/dataset/only_rpin/dataset_{fold}/{type}'
        else:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/fire/only_frequency/dataset/only_rpin/dataset_{fold}/{type}'
    elif samplingType==2:
        if nodeVecType == 0:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/frequency/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/pyfeat/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/pyfeat/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/pyfeat/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_n2v/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_n2v/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_n2v/dataset/only_rpin/dataset_{fold}/{type}'
        else:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_frequency/dataset/only_rpin/dataset_{fold}/{type}'
    elif samplingType==3:
        if nodeVecType == 0:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/frequency/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/pyfeat/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/pyfeat/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/pyfeat/dataset/only_rpin/dataset_{fold}/{type}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_n2v/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_n2v/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_n2v/dataset/only_rpin/dataset_{fold}/{type}'
        else:
            if subgraph_type == 0:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_frequency/dataset/rpin_ppin_rrsn/dataset_{fold}/{type}'
            elif subgraph_type == 1:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_frequency/dataset/rpin_with_ppin_rrsn/dataset_{fold}/{type}'
            else:
                dataset_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_frequency/dataset/only_rpin/dataset_{fold}/{type}'

    return dataset_path
def generate_log_path(projectName,model_code,nodeVecType,subgraph_type,samplingType):
    if samplingType==0:
        if nodeVecType == 0:  # random fire
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/frequency/only_rpin/model_{model_code}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/pyfeat/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/pyfeat/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/pyfeat/only_rpin/model_{model_code}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_n2v/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_n2v/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_n2v/only_rpin/model_{model_code}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/random/only_frequency/only_rpin/model_{model_code}'
    elif samplingType==1:
        if nodeVecType == 0:  # random fire
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/frequency/only_rpin/model_{model_code}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/pyfeat/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/pyfeat/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/pyfeat/only_rpin/model_{model_code}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_n2v/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_n2v/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_n2v/only_rpin/model_{model_code}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/fire/only_frequency/only_rpin/model_{model_code}'
    elif samplingType == 2:
        if nodeVecType == 0:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/frequency/only_rpin/model_{model_code}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/pyfeat/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/pyfeat/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/pyfeat/only_rpin/model_{model_code}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_n2v/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_n2v/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_n2v/only_rpin/model_{model_code}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable/only_frequency/only_rpin/model_{model_code}'
    elif samplingType == 3:
        if nodeVecType == 0:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/frequency/only_rpin/model_{model_code}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/pyfeat/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/pyfeat/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/pyfeat/only_rpin/model_{model_code}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_n2v/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_n2v/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_n2v/only_rpin/model_{model_code}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_frequency/rpin_ppin_rrsn/model_{model_code}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_frequency/rpin_with_ppin_rrsn/model_{model_code}'
            else:
                log_saving_path = f'../../data/log/{projectName}/rpi_hgnn/reliable_random/only_frequency/only_rpin/model_{model_code}'
    return log_saving_path

def generate_model_path(projectName, model_code, nodeVecType, subgraph_type,fold,samplingType):
    if samplingType==0:
        if nodeVecType == 0:  # random fire
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/frequency/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/pyfeat/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/pyfeat/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/pyfeat/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_n2v/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_n2v/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_n2v/only_rpin/model_{model_code}/fold_{fold}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/random/only_frequency/only_rpin/model_{model_code}/fold_{fold}'
    elif samplingType==1:
        if nodeVecType == 0:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/frequency/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/pyfeat/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/pyfeat/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/pyfeat/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_n2v/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_n2v/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_n2v/only_rpin/model_{model_code}/fold_{fold}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/fire/only_frequency/only_rpin/model_{model_code}/fold_{fold}'
    elif samplingType == 2:
        if nodeVecType == 0:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/frequency/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/pyfeat/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/pyfeat/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/pyfeat/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_n2v/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_n2v/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_n2v/only_rpin/model_{model_code}/fold_{fold}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable/only_frequency/only_rpin/model_{model_code}/fold_{fold}'
    elif samplingType == 3:
        if nodeVecType == 0:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/frequency/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 1:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/pyfeat/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/pyfeat/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/pyfeat/only_rpin/model_{model_code}/fold_{fold}'
        elif nodeVecType == 2:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_n2v/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_n2v/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_n2v/only_rpin/model_{model_code}/fold_{fold}'
        else:
            if subgraph_type == 0:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_frequency/rpin_ppin_rrsn/model_{model_code}/fold_{fold}'
            elif subgraph_type == 1:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_frequency/rpin_with_ppin_rrsn/model_{model_code}/fold_{fold}'
            else:
                log_saving_path = f'../../data/model/{projectName}/rpi_hgnn/reliable_random/only_frequency/only_rpin/model_{model_code}/fold_{fold}'
    return log_saving_path
def generate_pre_dataset_path(projectName,samplingType):
    if samplingType==0:
        pre_dataset_path = f'../../data/{projectName}/rpi_hgnn/pre_dataset/random'
    elif samplingType==1:
        pre_dataset_path= f'../../data/{projectName}/rpi_hgnn/pre_dataset/fire'
    elif samplingType == 2:
        pre_dataset_path = f'../../data/{projectName}/rpi_hgnn/pre_dataset/reliable'
    elif samplingType == 3:
        pre_dataset_path = f'../../data/{projectName}/rpi_hgnn/pre_dataset/reliable_random'
    return pre_dataset_path
def generate_node_path(projectName,samplingType,nodeVecType):
    if samplingType==0:
        if nodeVecType == 0:
            node_path = f'../../data/{projectName}/rpi_hgnn/random/frequency'
        elif nodeVecType == 1:
            node_path = f'../../data/{projectName}/rpi_hgnn/random/pyfeat'
        elif nodeVecType == 2:
            node_path = f'../../data/{projectName}/rpi_hgnn/random/only_n2v'
        else:
            node_path = f'../../data/{projectName}/rpi_hgnn/random/only_frequency'
    elif samplingType==1:
        if nodeVecType == 0:
            node_path = f'../../data/{projectName}/rpi_hgnn/fire/frequency'
        elif nodeVecType == 1:
            node_path = f'../../data/{projectName}/rpi_hgnn/fire/pyfeat'
        elif nodeVecType == 2:
            node_path = f'../../data/{projectName}/rpi_hgnn/fire/only_n2v'
        else:
            node_path = f'../../data/{projectName}/rpi_hgnn/fire/only_frequency'
    elif samplingType == 2:
        if nodeVecType == 0:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable/frequency'
        elif nodeVecType == 1:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable/pyfeat'
        elif nodeVecType == 2:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_n2v'
        else:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable/only_frequency'
    elif samplingType == 3:
        if nodeVecType == 0:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/frequency'
        elif nodeVecType == 1:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/pyfeat'
        elif nodeVecType == 2:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_n2v'
        else:
            node_path = f'../../data/{projectName}/rpi_hgnn/reliable_random/only_frequency'
    return node_path
def get_positive_samples_of_NPInter(NPI_filepath):
    NPInter = pd.read_table(NPI_filepath)
    protein = NPInter['Protein names'].unique().tolist()
    ncRNA = NPInter['RNA names'].unique().tolist()  #
    positive_index = []
    for index, row in NPInter.iterrows():
        i = ncRNA.index(row['RNA names'])
        j = protein.index(row['Protein names'])
        positive_index.append([i, j])
    return positive_index, protein, ncRNA
def calculate_rna_protein_similatity(rna, protein, positive_samples,swscore_matrix,dict_name_id):

    score = 0
    related_pair = [pair for pair in positive_samples if pair[0] == rna]
    protein_id = dict_name_id[protein]
    for pair in related_pair:
        if(pair[1]!= protein):
            pair1_id=dict_name_id[pair[1]]
            pp_sw=swscore_matrix[protein_id, pair1_id]/math.sqrt(swscore_matrix[protein_id, protein_id]*swscore_matrix[pair1_id, pair1_id])
            score += pp_sw
    return score
def calculate_rna_protein_similatity_plus(rna, protein, positive_samples,pp_swscore_matrix,dict_protein_name_id,rr_swscore_matrix,dict_rna_name_id):

    protein_score = 0
    rna_related_pair = [pair for pair in positive_samples if pair[0] == rna]
    protein_id = dict_protein_name_id[protein]
    protein_count=0
    for pair in rna_related_pair:
        if(pair[1]!= protein):
            pair1_id=dict_protein_name_id[pair[1]]
            pp_sw=pp_swscore_matrix[protein_id, pair1_id]/math.sqrt(pp_swscore_matrix[protein_id, protein_id]*pp_swscore_matrix[pair1_id, pair1_id])
            protein_score += pp_sw
            protein_count+=1
    rna_score = 0
    protein_related_pair = [pair for pair in positive_samples if pair[1] == protein]
    rna_id = dict_rna_name_id[rna]
    rna_count=0
    for pair in protein_related_pair:
        if(pair[0]!= rna):
            pair0_id=dict_rna_name_id[pair[0]]
            rr_sw=rr_swscore_matrix[rna_id, pair0_id]/math.sqrt(rr_swscore_matrix[rna_id, rna_id]*rr_swscore_matrix[pair0_id, pair0_id])
            rna_score += rr_sw
            rna_count+=1
    return protein_score,rna_score
def fire_negative_sampling (positive_samples, RNA_list, protein_list,swscore_matrix,dict_name_id,size):
    Positives = []
    Negatives = []
    for rna in RNA_list:
        for protein in protein_list:
            sample = [rna, protein]
            if [rna, protein] in positive_samples:
                Ms = 1
                sample.append(Ms)
                Positives.append(tuple(sample))
            else:
                Ms = calculate_rna_protein_similatity(rna, protein, positive_samples,swscore_matrix,dict_name_id)
                sample.append(Ms)
                Negatives.append(tuple(sample))
    Negatives = sorted(Negatives, key=lambda x: x[2])
    Negatives=Negatives[:size]
    return Positives, Negatives
def reliable_negative_sampling (positive_samples, RNA_list, protein_list,pp_swscore_matrix,dict_protein_name_id,rr_swscore_matrix,dict_rna_name_id,ratio,size):
    Positives = []
    Negatives = []
    for rna in RNA_list:
        for protein in protein_list:
            sample = [rna, protein]
            if [rna, protein] in positive_samples:
                Ms = 1
                sample.append(Ms)
                Positives.append(tuple(sample))
            else:
                protein_score,rna_score = calculate_rna_protein_similatity_plus(rna, protein, positive_samples,pp_swscore_matrix,dict_protein_name_id,rr_swscore_matrix,dict_rna_name_id)
                Ms=ratio*protein_score+(1-ratio)*rna_score
                sample.append(Ms)
                Negatives.append(tuple(sample))
    Negatives = sorted(Negatives, key=lambda x: x[2])
    Negatives=Negatives[:size]
    return Positives, Negatives
def case_study_sampling (positive_samples, RNA_list, protein_list,pp_swscore_matrix,dict_protein_name_id,rr_swscore_matrix,dict_rna_name_id,ratio,size):
    case_study_edges = []
    for rna in RNA_list:
        for protein in protein_list:
            sample = [rna, protein]
            if [rna, protein] not in positive_samples:
                protein_score,rna_score = calculate_rna_protein_similatity_plus(rna, protein, positive_samples,pp_swscore_matrix,dict_protein_name_id,rr_swscore_matrix,dict_rna_name_id)
                Ms=ratio*protein_score+(1-ratio)*rna_score
                sample.append(Ms)
                case_study_edges.append(tuple(sample))
    case_study_edges = sorted(case_study_edges, key=lambda x: x[2],reverse=True)
    case_study_edges=case_study_edges[:size]
    return case_study_edges